"""Validate the body-metric workbook and synchronize it to Supabase."""

from __future__ import annotations

import unicodedata
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from .db import upsert_rows, validate_supabase_config


EXPECTED_HEADERS = ("Index", "Measurement name", "Dimension", "Category")


def _clean_text(value: object, *, field: str, row_number: int) -> str:
    if not isinstance(value, str):
        raise ValueError(f"Row {row_number}: {field} must be text.")
    text = unicodedata.normalize("NFC", value).strip()
    if not text:
        raise ValueError(f"Row {row_number}: {field} cannot be blank.")
    if any(unicodedata.category(character) in {"Cc", "Cs"} for character in text):
        raise ValueError(f"Row {row_number}: {field} contains unsupported control characters.")
    return text


def read_metric_definitions(path: str | Path) -> list[dict[str, object]]:
    workbook_path = Path(path).expanduser().resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Metric workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) != 1:
            raise ValueError("The metric workbook must contain exactly one worksheet.")
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None or tuple(header[:4]) != EXPECTED_HEADERS:
            raise ValueError(f"Expected workbook headers: {', '.join(EXPECTED_HEADERS)}.")
        if any(value not in {None, ""} for value in header[4:]):
            raise ValueError("The metric workbook contains unexpected columns.")

        records: list[dict[str, object]] = []
        seen_ids: set[int] = set()
        seen_names: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            if not row or all(value in {None, ""} for value in row):
                continue
            if any(value not in {None, ""} for value in row[4:]):
                raise ValueError(f"Row {row_number}: unexpected data after Category.")
            raw_id = row[0]
            if isinstance(raw_id, bool) or not isinstance(raw_id, (int, float)) or int(raw_id) != raw_id:
                raise ValueError(f"Row {row_number}: Index must be a whole number.")
            metric_id = int(raw_id)
            if metric_id < 1 or metric_id in seen_ids:
                raise ValueError(f"Row {row_number}: Index must be positive and unique.")
            name = _clean_text(row[1], field="Measurement name", row_number=row_number)
            normalized_name = name.casefold()
            if normalized_name in seen_names:
                raise ValueError(f"Row {row_number}: Measurement name must be unique.")

            seen_ids.add(metric_id)
            seen_names.add(normalized_name)
            records.append(
                {
                    "id": metric_id,
                    "name": name,
                    "dimension": _clean_text(row[2], field="Dimension", row_number=row_number),
                    "category": _clean_text(row[3], field="Category", row_number=row_number),
                }
            )
    finally:
        workbook.close()

    if not records:
        raise ValueError("The metric workbook does not contain any metric rows.")
    return records


def sync_metrics_from_xlsx(path: str | Path, *, batch_size: int = 100) -> int:
    if batch_size < 1 or batch_size > 1000:
        raise ValueError("batch_size must be between 1 and 1000.")
    validate_supabase_config()
    records = read_metric_definitions(path)
    updated_at = datetime.now(UTC).isoformat()
    rows = [{**record, "updated_at": updated_at} for record in records]
    try:
        for start in range(0, len(rows), batch_size):
            upsert_rows("body_metrics", rows[start : start + batch_size], on_conflict="id")
    except Exception as exc:
        if getattr(exc, "code", None) in {"42P01", "PGRST205"}:
            raise RuntimeError(
                "Supabase table public.body_metrics does not exist. Run the SQL from "
                "GET /api/metrics/schema in the Supabase SQL Editor, then retry."
            ) from exc
        raise
    return len(rows)
